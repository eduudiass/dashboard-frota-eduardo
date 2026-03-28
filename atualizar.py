"""
FleetManager — Gerador de dados.json
Duplo clique ou rode: python atualizar.py
Lê a planilha e atualiza o dados.json no repositório.
"""

import json, datetime, sys, os

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# ═══ CAMINHOS ═══
PLANILHA = r"C:\Users\Eduardo Dias\Downloads\Eduardo's Transportes\Gestao_Frota_Atual.xlsx"
REPO = r"C:\Users\Eduardo Dias\Downloads\dashboard-frota-eduardo"
OUTPUT = os.path.join(REPO, "dados.json")

if not os.path.exists(PLANILHA):
    print(f"ERRO: Planilha nao encontrada em:\n  {PLANILHA}")
    input("Pressione Enter para sair...")
    sys.exit(1)

print(f"Lendo: {PLANILHA}")

wb = openpyxl.load_workbook(PLANILHA, data_only=True)

def date_str(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%d/%m/%Y')
    return str(v) if v else ''

def safe_str(v):
    if v is None or v == 0:
        return ''
    return str(v)

def safe_float(v):
    try:
        return float(v) if v else 0.0
    except:
        return 0.0

def safe_int(v):
    try:
        return int(v) if v else 0
    except:
        return 0

data = {
    "veiculos": [],
    "manutencao_resumo": [],
    "manutencao_registros": [],
    "receitas": [],
    "gastos": [],
    "compras": [],
    "vendas_resumo": {},
    "vendas_registros": [],
    "resumo": {}
}

# ═══ VEICULOS ═══
ws = wb['Cadastro de Veículos']
for r in range(4, 50):
    n = ws.cell(r, 1).value
    placa = ws.cell(r, 2).value
    if not n or not placa:
        continue
    if str(n).upper().startswith('TOTAL'):
        break
    data['veiculos'].append({
        "numero": safe_int(n),
        "placa": str(placa),
        "modelo": safe_str(ws.cell(r, 3).value),
        "ano": safe_int(ws.cell(r, 4).value),
        "cor": safe_str(ws.cell(r, 5).value),
        "motorista": safe_str(ws.cell(r, 6).value) or 'Sem motorista',
        "contato": safe_str(ws.cell(r, 7).value),
        "aluguel_mensal": safe_float(ws.cell(r, 8).value),
        "data_inicio": date_str(ws.cell(r, 9).value),
        "km_atual": safe_float(ws.cell(r, 10).value),
        "status": safe_str(ws.cell(r, 11).value) or 'Ativo'
    })

print(f"  Veiculos: {len(data['veiculos'])}")

# ═══ MANUTENCAO RESUMO ═══
ws = wb['Manutenção']
for r in range(4, 50):
    n = ws.cell(r, 1).value
    placa = ws.cell(r, 2).value
    if str(n).upper().startswith('TOTAL') if n else False:
        break
    if not n or not placa or placa == 0:
        continue
    data['manutencao_resumo'].append({
        "numero": safe_int(n),
        "placa": str(placa),
        "modelo": safe_str(ws.cell(r, 3).value),
        "motorista": safe_str(ws.cell(r, 4).value) or 'Sem motorista',
        "num_servicos": safe_int(ws.cell(r, 5).value),
        "total_gasto": safe_float(ws.cell(r, 6).value),
        "status": safe_str(ws.cell(r, 10).value) or 'Ativo'
    })

# ═══ MANUTENCAO REGISTROS ═══
reg_start = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v and 'REGISTROS DE MANUTEN' in str(v).upper():
        reg_start = r + 2
        break

if reg_start:
    for r in range(reg_start, ws.max_row + 1):
        n = ws.cell(r, 1).value
        placa = ws.cell(r, 3).value
        if not n or not placa:
            continue
        if not isinstance(n, (int, float)):
            continue
        data['manutencao_registros'].append({
            "numero": safe_int(n),
            "data": date_str(ws.cell(r, 2).value),
            "placa": str(placa),
            "modelo": safe_str(ws.cell(r, 4).value),
            "tipo": safe_str(ws.cell(r, 5).value),
            "descricao": safe_str(ws.cell(r, 6).value),
            "oficina": safe_str(ws.cell(r, 7).value),
            "custo": safe_float(ws.cell(r, 8).value),
            "km_servico": safe_str(ws.cell(r, 9).value),
            "prox_revisao": safe_str(ws.cell(r, 10).value)
        })

print(f"  Manutencoes: {len(data['manutencao_resumo'])} resumos, {len(data['manutencao_registros'])} registros")

# ═══ RECEITAS ═══
ws = wb['Receitas Mensais']
meses_cols = {'jan':6,'fev':7,'mar':8,'abr':9,'mai':10,'jun':11,'jul':12,'ago':13,'set':14,'out':15,'nov':16,'dez':17}
for r in range(4, 50):
    n = ws.cell(r, 1).value
    placa = ws.cell(r, 2).value
    if not n or not placa:
        continue
    if str(n).upper().startswith('TOTAL'):
        break
    rec = {}
    for mes, col in meses_cols.items():
        rec[mes] = safe_float(ws.cell(r, col).value)
    data['receitas'].append({
        "numero": safe_int(n),
        "placa": str(placa),
        "modelo": safe_str(ws.cell(r, 3).value),
        "motorista": safe_str(ws.cell(r, 4).value) or 'Sem motorista',
        "aluguel_previsto": safe_float(ws.cell(r, 5).value),
        "receitas_mensais": rec,
        "total_recebido": safe_float(ws.cell(r, 18).value),
        "pct_recebido": safe_float(ws.cell(r, 19).value)
    })

print(f"  Receitas: {len(data['receitas'])}")

# ═══ GASTOS ═══
ws = wb['Gastos']
for r in range(4, 50):
    n = ws.cell(r, 1).value
    placa = ws.cell(r, 2).value
    if not n or not placa:
        continue
    if str(n).upper().startswith('TOTAL'):
        break
    data['gastos'].append({
        "numero": safe_int(n),
        "placa": str(placa),
        "modelo": safe_str(ws.cell(r, 3).value),
        "motorista": safe_str(ws.cell(r, 4).value) or 'Sem motorista',
        "seguro_anual": safe_float(ws.cell(r, 5).value),
        "ipva_anual": safe_float(ws.cell(r, 6).value),
        "licenciamento": safe_float(ws.cell(r, 7).value),
        "combustivel_mensal": safe_float(ws.cell(r, 8).value),
        "total_mensal": safe_float(ws.cell(r, 9).value)
    })

print(f"  Gastos: {len(data['gastos'])}")

# ═══ COMPRAS ═══
ws = wb['🚗 Compra de Carros']
carro = None
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a and isinstance(a, str) and '🚗' in a and '—' in a:
        if carro:
            data['compras'].append(carro)
        carro = {"titulo": a.strip(), "itens": [], "total": 0}
        continue
    if a and isinstance(a, str) and a.upper() == 'TOTAL':
        if carro:
            carro['total'] = safe_float(ws.cell(r, 5).value)
        continue
    if carro and a and isinstance(a, (int, float)) and a > 0:
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value
        if c or d or e:
            carro['itens'].append({
                "numero": safe_int(a),
                "data": date_str(ws.cell(r, 2).value),
                "categoria": safe_str(c),
                "descricao": safe_str(d),
                "valor": safe_float(e),
                "observacao": safe_str(ws.cell(r, 6).value)
            })
if carro:
    data['compras'].append(carro)
data['compras'] = [c for c in data['compras'] if c['total'] > 0 or len(c['itens']) > 0]

print(f"  Compras: {len(data['compras'])} carros")

# ═══ VENDAS ═══
if '🚗 Vendas' in wb.sheetnames:
    ws = wb['🚗 Vendas']
    data['vendas_resumo'] = {
        "carros_vendidos": safe_int(ws.cell(6, 1).value),
        "total_investido": safe_float(ws.cell(6, 3).value),
        "total_arrecadado": safe_float(ws.cell(6, 5).value),
        "lucro_total": safe_float(ws.cell(6, 7).value),
        "lucro_medio": safe_float(ws.cell(6, 9).value)
    }
    for r in range(9, 60):
        n = ws.cell(r, 1).value
        if not n:
            continue
        placa = ws.cell(r, 3).value
        modelo = ws.cell(r, 4).value
        venda = ws.cell(r, 7).value
        if placa or modelo or venda:
            data['vendas_registros'].append({
                "numero": safe_int(n),
                "data_venda": date_str(ws.cell(r, 2).value),
                "placa": safe_str(placa),
                "modelo": safe_str(modelo),
                "preco_compra": safe_float(ws.cell(r, 5).value),
                "gastos_mecanica": safe_float(ws.cell(r, 6).value),
                "valor_venda": safe_float(venda),
                "custo_total": safe_float(ws.cell(r, 8).value),
                "lucro": safe_float(ws.cell(r, 9).value),
                "margem": safe_float(ws.cell(r, 10).value)
            })

    print(f"  Vendas: {len(data['vendas_registros'])} registros")

# ═══ RESUMO ═══
res = data['resumo']
res['receita_mensal'] = sum(v['aluguel_mensal'] for v in data['veiculos'])
res['total_manutencao'] = sum(m['total_gasto'] for m in data['manutencao_resumo'])
res['num_servicos'] = sum(m['num_servicos'] for m in data['manutencao_resumo'])
res['gastos_fixos_mensal'] = sum(g['total_mensal'] for g in data['gastos'])
res['total_investido_compras'] = sum(c['total'] for c in data['compras'])
res['total_veiculos'] = len(data['veiculos'])
res['veiculos_ativos'] = sum(1 for v in data['veiculos'] if v['status'] == 'Ativo')
res['veiculos_inativos'] = res['total_veiculos'] - res['veiculos_ativos']
res['total_combustivel_mes'] = sum(g['combustivel_mensal'] for g in data['gastos'])
res['total_km'] = sum(v['km_atual'] for v in data['veiculos'])
res['media_por_carro'] = res['receita_mensal'] / res['total_veiculos'] if res['total_veiculos'] > 0 else 0
res['lucro_mensal'] = res['receita_mensal'] - res['gastos_fixos_mensal'] - res['total_manutencao']

# ═══ SALVAR ═══
with open(OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"\n✅ dados.json atualizado!")
print(f"   Salvo em: {OUTPUT}")
input("\nPressione Enter para sair...")
